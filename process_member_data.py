#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
乐读会员收入数据自动处理脚本
完整流程:
1. 从Downloads找到最新的"数据表 (n).csv"
2. 清洗数据(过滤PAID、标准化商品名、拆分日期时间) → 第1个sheet
3. 按【渠道】分流:
   - 包含 app/default → 【APP内】sheet
   - 其他 → 【外部渠道】sheet
4. 【统计】sheet的公式自动计算
"""

import pandas as pd
import os
import re
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

def find_latest_data_file(download_dir):
    """在Downloads目录中查找最新的数据表文件"""
    pattern = re.compile(r'数据表\s*\((\d+)\)\.csv')
    max_num = -1
    latest_file = None

    for file in os.listdir(download_dir):
        match = pattern.match(file)
        if match:
            num = int(match.group(1))
            if num > max_num:
                max_num = num
                latest_file = os.path.join(download_dir, file)

    return latest_file

def clean_data(source_file):
    """
    第一步:清洗原始数据
    - 过滤PAID订单
    - 标准化商品名称
    - 拆分支付时间为日期和时间
    """
    print(f"\n{'='*50}")
    print(f"步骤1: 清洗原始数据")
    print(f"{'='*50}")
    print(f"源文件: {source_file}")

    # 读取源数据
    df = pd.read_csv(source_file)
    print(f"✓ 读取源数据: {len(df)} 行")

    # 过滤:只保留已支付订单
    df = df[df['支付状态'] == 'PAID'].copy()
    print(f"✓ 过滤后(仅PAID): {len(df)} 行")

    # 标准化商品名称
    if '商品名称' in df.columns:
        name_mapping = {
            '30天': '30天乐读会员',
            '90天': '90天乐读会员',
            '365天': '365天乐读会员'
        }
        original_counts = df['商品名称'].value_counts()
        df['商品名称'] = df['商品名称'].replace(name_mapping)

        standardized = []
        for old_name, new_name in name_mapping.items():
            if old_name in original_counts:
                count = original_counts[old_name]
                standardized.append(f"'{old_name}' → '{new_name}' ({count}条)")

        if standardized:
            print(f"✓ 商品名称标准化: {', '.join(standardized)}")

    # 处理下单时间,拆分为日期和时间
    if '下单时间' in df.columns:
        df['下单时间'] = df['下单时间'].astype(str).str.strip()
        df['下单时间_dt'] = pd.to_datetime(df['下单时间'], errors='coerce')
        df['日期'] = df['下单时间_dt'].dt.strftime('%Y-%m-%d')
        df['时间'] = df['下单时间_dt'].dt.strftime('%H:%M:%S')
        df = df.drop('下单时间_dt', axis=1)

        # 调整列顺序:将【日期】【时间】插入到【下单时间】之后
        cols = df.columns.tolist()
        if '日期' in cols and '时间' in cols and '下单时间' in cols:
            cols.remove('日期')
            cols.remove('时间')
            order_time_idx = cols.index('下单时间')
            cols.insert(order_time_idx + 1, '日期')
            cols.insert(order_time_idx + 2, '时间')
            df = df[cols]

        print(f"✓ 拆分下单时间为 [日期] 和 [时间]")

    return df

def split_by_channel(df):
    """
    第二步:按【渠道】分流数据
    - 包含 app 或 default → APP内
    - 其他 → 外部渠道
    """
    print(f"\n{'='*50}")
    print(f"步骤2: 按【渠道】分流数据")
    print(f"{'='*50}")

    if '渠道' not in df.columns:
        print("❌ 警告: 未找到【渠道】列")
        return df, pd.DataFrame()

    # 筛选包含 app 或 default 的渠道
    mask_app = df['渠道'].astype(str).str.lower().str.contains('app|default', case=False, na=False)
    df_app = df[mask_app].copy()
    df_external = df[~mask_app].copy()

    print(f"✓ 【APP内】: {len(df_app)} 行 (渠道包含 'app' 或 'default')")
    print(f"✓ 【外部渠道】: {len(df_external)} 行 (其他所有渠道)")

    return df_app, df_external

def save_to_excel(df_all, df_app, df_external, target_file):
    """
    第三步:保存到Excel的4个sheet
    - Sheet1: 全部清洗后的数据
    - APP内: APP和default渠道的数据
    - 外部渠道: 其他渠道的数据
    - 统计: 保留现有公式(如果存在)
    """
    print(f"\n{'='*50}")
    print(f"步骤3: 保存到Excel (4个sheet)")
    print(f"{'='*50}")

    # 检查目标文件是否存在
    has_existing_stats = False
    stats_sheet = None

    if os.path.exists(target_file):
        try:
            wb_existing = load_workbook(target_file)
            if '统计' in wb_existing.sheetnames:
                has_existing_stats = True
                stats_sheet = wb_existing['统计']
                print(f"✓ 检测到现有【统计】sheet,将保留其公式")
            wb_existing.close()
        except Exception as e:
            print(f"  警告: 无法读取现有文件 - {e}")

    # 使用pandas写入数据到临时文件
    with pd.ExcelWriter(target_file, engine='openpyxl', mode='w') as writer:
        df_all.to_excel(writer, sheet_name='Sheet1', index=False)
        df_app.to_excel(writer, sheet_name='APP内', index=False)
        df_external.to_excel(writer, sheet_name='外部渠道', index=False)

        # 如果没有现有的统计sheet,创建一个空的
        if not has_existing_stats:
            pd.DataFrame().to_excel(writer, sheet_name='统计', index=False)

    print(f"✓ 保存到: {target_file}")
    print(f"  - Sheet1: {len(df_all)} 行 (全部数据)")
    print(f"  - APP内: {len(df_app)} 行")
    print(f"  - 外部渠道: {len(df_external)} 行")
    print(f"  - 统计: {'保留现有公式' if has_existing_stats else '已创建空sheet'}")

    # 如果有现有的统计sheet,复制回去
    if has_existing_stats and stats_sheet:
        try:
            wb = load_workbook(target_file)
            if '统计' in wb.sheetnames:
                del wb['统计']
            wb.create_sheet('统计')
            new_stats = wb['统计']

            # 复制单元格内容和公式
            for row in stats_sheet.iter_rows():
                for cell in row:
                    new_cell = new_stats[cell.coordinate]
                    new_cell.value = cell.value
                    if cell.has_style:
                        new_cell.font = cell.font.copy()
                        new_cell.border = cell.border.copy()
                        new_cell.fill = cell.fill.copy()
                        new_cell.number_format = cell.number_format
                        new_cell.protection = cell.protection.copy()
                        new_cell.alignment = cell.alignment.copy()

            wb.save(target_file)
            wb.close()
            print(f"✓ 【统计】sheet 的公式已保留并更新")
        except Exception as e:
            print(f"  警告: 复制统计sheet时出错 - {e}")

def process_member_data(source_file, target_file):
    """主处理流程"""
    print(f"\n{'='*60}")
    print(f"🚀 乐读会员收入数据自动处理")
    print(f"{'='*60}")
    print(f"开始时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    try:
        # 第一步:清洗数据
        df_all = clean_data(source_file)

        # 第二步:按渠道分流
        df_app, df_external = split_by_channel(df_all)

        # 第三步:保存到Excel
        save_to_excel(df_all, df_app, df_external, target_file)

        print(f"\n{'='*60}")
        print(f"✅ 处理完成!")
        print(f"结束时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"{'='*60}\n")

        return len(df_all)

    except Exception as e:
        print(f"\n❌ 处理失败: {str(e)}")
        import traceback
        traceback.print_exc()
        raise

def main():
    """主函数"""
    # 定义路径
    download_dir = os.path.expanduser('~/Downloads')
    target_file = os.path.expanduser('~/Desktop/乐读会员收入.xlsx')

    try:
        # 查找最新的数据文件
        source_file = find_latest_data_file(download_dir)

        if not source_file:
            print(f"错误: 在 {download_dir} 中未找到数据表文件")
            print("请确保文件名格式为: 数据表 (n).csv")
            return 1

        # 处理数据
        row_count = process_member_data(source_file, target_file)

        print(f"\n✅ 成功处理 {row_count} 条会员订单数据")
        return 0

    except Exception as e:
        print(f"\n❌ 处理失败: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    exit(main())
